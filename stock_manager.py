#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
주식 포트폴리오 매니저
- 한국투자증권(KIS) OpenAPI 또는 pykrx(무료) 사용
- 보유 종목 현재가, 평가손익, 수익률 등 자동 계산
- 종목 코드/종목명 검색 지원
"""

import json
import os
import sys
import requests
from datetime import datetime, date, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings("ignore")

# ============================================================
#  상수 / 설정
# ============================================================
CONFIG_FILE = "config.json"
EXCEL_FILE  = "주식_포트폴리오.xlsx"

DEFAULT_CONFIG = {
    "api_provider": "pykrx",        # "pykrx" 또는 "kis"
    "kis": {
        "app_key":    "YOUR_APP_KEY",
        "app_secret": "YOUR_APP_SECRET",
        "account_no": "YOUR_ACCOUNT_NO",  # 예) "50123456-01" 또는 "5012345601"
        "is_paper":   False               # True: 모의투자, False: 실전
    },
    "portfolio": [
        # KIS API 미사용 시 수동 입력 (또는 엑셀 '보유종목 입력' 시트 사용)
        # {"code": "005930", "quantity": 10, "avg_price": 72000},
    ]
}

# ---- 색상 팔레트 ----
C_HEADER_BG   = "1F3864"   # 헤더 배경 (네이비)
C_HEADER_FG   = "FFFFFF"   # 헤더 글자 (흰색)
C_ROW_ODD     = "F0F5FF"   # 홀수행
C_ROW_EVEN    = "FFFFFF"   # 짝수행
C_UP          = "C00000"   # 상승 (빨강)
C_DOWN        = "0070C0"   # 하락 (파랑)
C_TOTAL_BG    = "FFF2CC"   # 합계행
C_SECTION_BG  = "D6E4F7"   # 섹션/소제목
C_INPUT_BG    = "FFFFF0"   # 입력셀
C_NOTE_BG     = "FFFDE7"   # 안내문

# ============================================================
#  KIS OpenAPI 클라이언트
# ============================================================
class KISApi:
    REAL_URL  = "https://openapi.kis.or.kr"
    PAPER_URL = "https://openapivts.koreainvestment.com:29443"

    def __init__(self, app_key, app_secret, account_no, is_paper=False):
        self.app_key    = app_key
        self.app_secret = app_secret
        raw = account_no.replace("-", "")
        self.cano          = raw[:8]
        self.acnt_prdt_cd  = raw[8:] if len(raw) > 8 else "01"
        self.base_url      = self.PAPER_URL if is_paper else self.REAL_URL
        self.is_paper      = is_paper
        self.access_token  = None

    def authenticate(self):
        try:
            resp = requests.post(
                f"{self.base_url}/oauth2/tokenP",
                json={"grant_type": "client_credentials",
                      "appkey": self.app_key, "appsecret": self.app_secret},
                timeout=10
            )
            resp.raise_for_status()
            self.access_token = resp.json().get("access_token", "")
            return bool(self.access_token)
        except Exception as e:
            print(f"  [KIS] 인증 오류: {e}")
            return False

    def _headers(self, tr_id):
        return {
            "content-type":  "application/json; charset=utf-8",
            "authorization": f"Bearer {self.access_token}",
            "appkey":        self.app_key,
            "appsecret":     self.app_secret,
            "tr_id":         tr_id,
            "custtype":      "P",
        }

    def get_stock_price(self, code):
        """현재가 조회 → dict 또는 None"""
        try:
            resp = requests.get(
                f"{self.base_url}/uapi/domestic-stock/v1/quotations/inquire-price",
                headers=self._headers("FHKST01010100"),
                params={"FID_COND_MRKT_DIV_CODE": "J", "FID_INPUT_ISCD": code},
                timeout=10,
            )
            resp.raise_for_status()
            d = resp.json()
            if d.get("rt_cd") == "0":
                o = d["output"]
                return {
                    "code":          code,
                    "name":          o.get("hts_kor_isnm", ""),
                    "current_price": int(o.get("stck_prpr", 0)),
                    "prev_price":    int(o.get("stck_sdpr", 0)),
                    "change":        int(o.get("prdy_vrss", 0)),
                    "change_rate":   float(o.get("prdy_ctrt", 0)),
                    "volume":        int(o.get("acml_vol", 0)),
                    "open":          int(o.get("stck_oprc", 0)),
                    "high":          int(o.get("stck_hgpr", 0)),
                    "low":           int(o.get("stck_lwpr", 0)),
                }
        except Exception as e:
            print(f"  [KIS] 현재가 조회 실패 ({code}): {e}")
        return None

    def get_balance(self):
        """잔고 조회 → dict 또는 None"""
        try:
            tr_id = "VTTC8434R" if self.is_paper else "TTTC8434R"
            resp = requests.get(
                f"{self.base_url}/uapi/domestic-stock/v1/trading/inquire-balance",
                headers=self._headers(tr_id),
                params={
                    "CANO": self.cano, "ACNT_PRDT_CD": self.acnt_prdt_cd,
                    "AFHR_FLPR_YN": "N", "OFL_YN": "", "INQR_DVSN": "02",
                    "UNPR_DVSN": "01", "FUND_STTL_ICLD_YN": "N",
                    "FNCG_AMT_AUTO_RDPT_YN": "N", "PRCS_DVSN": "01",
                    "CTX_AREA_FK100": "", "CTX_AREA_NK100": "",
                },
                timeout=10,
            )
            resp.raise_for_status()
            d = resp.json()
            if d.get("rt_cd") == "0":
                holdings = []
                for item in d.get("output1", []):
                    qty = int(item.get("hldg_qty", 0))
                    if qty > 0:
                        holdings.append({
                            "code":          item.get("pdno", ""),
                            "name":          item.get("prdt_name", ""),
                            "quantity":      qty,
                            "avg_price":     float(item.get("pchs_avg_pric", 0)),
                            "current_price": int(item.get("prpr", 0)),
                            "eval_amount":   int(item.get("evlu_amt", 0)),
                            "buy_amount":    int(item.get("pchs_amt", 0)),
                            "profit_loss":   int(item.get("evlu_pfls_amt", 0)),
                            "profit_rate":   float(item.get("evlu_pfls_rt", 0)),
                            "change":        0,
                            "change_rate":   0.0,
                        })
                s = (d.get("output2") or [{}])[0]
                return {
                    "holdings":          holdings,
                    "total_eval":        int(s.get("scts_evlu_amt", 0)),
                    "total_buy":         int(s.get("pchs_amt_smtl_amt", 0)),
                    "total_profit":      int(s.get("evlu_pfls_smtl_amt", 0)),
                    "total_profit_rate": float(s.get("evlu_erng_rt", 0)),
                    "deposit":           int(s.get("dnca_tot_amt", 0)),
                    "total_asset":       int(s.get("tot_evlu_amt", 0)),
                }
        except Exception as e:
            print(f"  [KIS] 잔고 조회 실패: {e}")
        return None


# ============================================================
#  네이버 금융 공개 API (pykrx 미설치 시 폴백)
# ============================================================
class NaverStockProvider:
    """네이버 금융 모바일 API (인증 불필요, numpy 의존성 없음)"""

    BASE  = "https://m.stock.naver.com/api/stock"
    SRCH  = "https://ac.finance.naver.com/ac"

    def __init__(self):
        self.available = True
        self._headers  = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                          "AppleWebKit/537.36 (KHTML, like Gecko) "
                          "Chrome/120.0.0.0 Safari/537.36",
            "Referer":    "https://m.stock.naver.com/",
        }

    def get_stock_price(self, code):
        """현재가 조회 → dict 또는 None"""
        try:
            resp = requests.get(
                f"{self.BASE}/{code}/basic",
                headers=self._headers, timeout=10, verify=False
            )
            if resp.status_code != 200:
                return None
            d = resp.json()

            current  = int(str(d.get("closePrice", "0")).replace(",", "") or 0)
            change   = int(str(d.get("compareToPreviousClosePrice", "0")).replace(",", "") or 0)
            cr_str   = str(d.get("fluctuationsRatio", "0")).replace("%", "").replace(",", "")
            try:
                change_rate = float(cr_str)
            except ValueError:
                change_rate = 0.0

            # 상승/하락 부호 보정
            sign = d.get("compareToPreviousPrice", {})
            if isinstance(sign, dict):
                code_sign = sign.get("code", "")
                if code_sign in ("2", "하락"):
                    change      = -abs(change)
                    change_rate = -abs(change_rate)

            return {
                "code":          code,
                "name":          d.get("stockName", code),
                "current_price": current,
                "prev_price":    current - change,
                "change":        change,
                "change_rate":   change_rate,
                "volume":        int(str(d.get("accumulatedTradingVolume","0")).replace(",","") or 0),
                "open":          int(str(d.get("openingPrice","0")).replace(",","") or 0),
                "high":          int(str(d.get("highPrice","0")).replace(",","") or 0),
                "low":           int(str(d.get("lowPrice","0")).replace(",","") or 0),
            }
        except Exception as e:
            print(f"  [Naver] {code} 조회 실패: {e}")
        return None

    def search_by_name(self, keyword):
        """종목명 자동완성 검색"""
        try:
            resp = requests.get(
                self.SRCH,
                params={"query": keyword, "target": "stock,index,marketindicator"},
                headers=self._headers, timeout=10, verify=False
            )
            if resp.status_code != 200:
                return []
            items = resp.json().get("items", [[]])
            results = []
            for group in items:
                for item in group:
                    if len(item) >= 2:
                        results.append({"code": item[1], "name": item[0]})
            return results[:10]
        except Exception as e:
            print(f"  [Naver] 종목 검색 실패: {e}")
        return []


# ============================================================
#  pykrx 래퍼 (무료, 인증 불필요)
# ============================================================
class PykrxProvider:
    def __init__(self):
        try:
            from pykrx import stock
            self._m = stock
            self.available = True
            self._name_cache = {}
            print("  [pykrx] 로드 성공")
        except ImportError:
            self.available = False
            print("  [pykrx] 미설치 — pip install pykrx 를 실행하세요.")

    def _recent_date(self):
        """최근 영업일 YYYYMMDD 반환"""
        for i in range(10):
            d = (date.today() - timedelta(days=i)).strftime("%Y%m%d")
            try:
                df = self._m.get_market_ohlcv(d, d, "005930")
                if not df.empty:
                    return d
            except Exception:
                continue
        return date.today().strftime("%Y%m%d")

    def _prev_date(self, ref_yyyymmdd):
        """ref 이전 영업일 반환"""
        ref = datetime.strptime(ref_yyyymmdd, "%Y%m%d").date()
        for i in range(1, 10):
            d = (ref - timedelta(days=i)).strftime("%Y%m%d")
            try:
                df = self._m.get_market_ohlcv(d, d, "005930")
                if not df.empty:
                    return d
            except Exception:
                continue
        return ref_yyyymmdd

    def _ticker_name(self, code):
        if code not in self._name_cache:
            try:
                self._name_cache[code] = self._m.get_market_ticker_name(code)
            except Exception:
                self._name_cache[code] = code
        return self._name_cache[code]

    def get_stock_price(self, code):
        if not self.available:
            return None
        try:
            today_str = self._recent_date()
            prev_str  = self._prev_date(today_str)

            df_today = self._m.get_market_ohlcv(today_str, today_str, code)
            df_prev  = self._m.get_market_ohlcv(prev_str,  prev_str,  code)

            if df_today.empty:
                return None

            r = df_today.iloc[-1]
            # 컬럼명 한/영 호환
            def col(kr, en):
                return r[kr] if kr in r.index else r.get(en, 0)

            current   = int(col("종가", "Close"))
            prev_close = int(df_prev.iloc[-1][
                "종가" if "종가" in df_prev.columns else "Close"
            ]) if not df_prev.empty else current
            change      = current - prev_close
            change_rate = round(change / prev_close * 100, 2) if prev_close else 0.0

            return {
                "code":          code,
                "name":          self._ticker_name(code),
                "current_price": current,
                "prev_price":    prev_close,
                "change":        change,
                "change_rate":   change_rate,
                "volume":        int(col("거래량", "Volume")),
                "open":          int(col("시가",   "Open")),
                "high":          int(col("고가",   "High")),
                "low":           int(col("저가",   "Low")),
            }
        except Exception as e:
            print(f"  [pykrx] {code} 조회 실패: {e}")
        return None

    def search_by_name(self, keyword):
        """종목명 검색 → [{"code", "name"}, ...]"""
        if not self.available:
            return []
        try:
            result = []
            for market in ["KOSPI", "KOSDAQ"]:
                tickers = self._m.get_market_ticker_list(market=market)
                for t in tickers:
                    n = self._ticker_name(t)
                    if keyword in n:
                        result.append({"code": t, "name": n})
            return result[:10]
        except Exception as e:
            print(f"  [pykrx] 종목명 검색 실패: {e}")
        return []


# ============================================================
#  엑셀 유틸리티
# ============================================================
def thin_border(color="CCCCCC"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def header_cell(cell, text, font_size=10):
    cell.value = text
    cell.font      = Font(name="맑은 고딕", bold=True, size=font_size, color=C_HEADER_FG)
    cell.fill      = PatternFill("solid", fgColor=C_HEADER_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = thin_border("888888")

def data_cell(cell, value, fmt=None, align="center", row_idx=0, bold=False):
    cell.value     = value
    bg = C_ROW_ODD if row_idx % 2 == 0 else C_ROW_EVEN
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.font      = Font(name="맑은 고딕", size=10, bold=bold)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border    = thin_border()
    if fmt:
        cell.number_format = fmt

def color_pnl(cell, value):
    """양수=빨강, 음수=파랑, 0=검정"""
    bold = cell.font.bold
    sz   = cell.font.size or 10
    if value > 0:
        cell.font = Font(name="맑은 고딕", size=sz, bold=bold, color=C_UP)
    elif value < 0:
        cell.font = Font(name="맑은 고딕", size=sz, bold=bold, color=C_DOWN)


# ============================================================
#  시트 생성
# ============================================================
def build_portfolio_sheet(ws, portfolio_data):
    ws.title = "포트폴리오"

    # ── 제목 ──
    ws.merge_cells("A1:L1")
    c = ws["A1"]
    c.value     = f"📈  주식 포트폴리오 현황   |   업데이트: {datetime.now().strftime('%Y-%m-%d  %H:%M:%S')}"
    c.font      = Font(name="맑은 고딕", bold=True, size=13, color="1F3864")
    c.fill      = PatternFill("solid", fgColor=C_SECTION_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # ── 요약 박스 (행2) ──
    summary = portfolio_data.get("summary", {})
    total_eval   = summary.get("total_eval", 0)
    total_buy    = summary.get("total_buy", 0)
    total_profit = summary.get("total_profit", 0)
    total_prate  = summary.get("total_profit_rate", 0)

    def summary_label(cell_addr, text):
        c = ws[cell_addr]
        c.value     = text
        c.font      = Font(name="맑은 고딕", bold=True, size=9, color="444444")
        c.fill      = PatternFill("solid", fgColor="EEF4FF")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = thin_border("BBBBBB")

    def summary_value(cell_addr, value, fmt, color=None):
        c = ws[cell_addr]
        c.value          = value
        c.number_format  = fmt
        c.font           = Font(name="맑은 고딕", bold=True, size=11,
                                color=color or "1F3864")
        c.fill           = PatternFill("solid", fgColor="FFFFFF")
        c.alignment      = Alignment(horizontal="right", vertical="center")
        c.border         = thin_border("BBBBBB")

    pc = C_UP if total_profit >= 0 else C_DOWN
    pr = C_UP if total_prate >= 0 else C_DOWN

    ws.merge_cells("A2:B2");  summary_label("A2", "매입금액 합계")
    ws.merge_cells("C2:D2");  summary_value("C2", total_buy, '#,##0"원"')
    ws.merge_cells("E2:F2");  summary_label("E2", "평가금액 합계")
    ws.merge_cells("G2:H2");  summary_value("G2", total_eval, '#,##0"원"')
    ws.merge_cells("I2:J2");  summary_label("I2", "총 평가손익")
    ws.merge_cells("K2:L2");  summary_value("K2", total_profit, '+#,##0;-#,##0;0"원"', pc)
    ws.row_dimensions[2].height = 28

    ws.merge_cells("A3:D3");  summary_label("A3", "예수금")
    deposit = summary.get("deposit", 0)
    ws.merge_cells("E3:F3");  summary_value("E3", deposit if deposit else "-", '#,##0"원"')
    ws.merge_cells("G3:H3");  summary_label("G3", "총자산")
    total_asset = summary.get("total_asset", 0)
    ws.merge_cells("I3:J3");  summary_value("I3", total_asset if total_asset else "-", '#,##0"원"')
    ws.merge_cells("K3:L3")
    rate_c = ws["K3"]
    rate_c.value          = total_prate / 100 if total_prate else 0
    rate_c.number_format  = '+0.00%;-0.00%;0.00%'
    rate_c.font           = Font(name="맑은 고딕", bold=True, size=13, color=pr)
    rate_c.fill           = PatternFill("solid", fgColor="FFFFFF")
    rate_c.alignment      = Alignment(horizontal="center", vertical="center")
    rate_c.border         = thin_border("BBBBBB")
    ws.row_dimensions[3].height = 26

    ws.row_dimensions[4].height = 6   # 여백

    # ── 컬럼 헤더 (행5) ──
    cols = [
        ("종목코드", "A",  9),
        ("종목명",   "B", 16),
        ("보유수량", "C", 10),
        ("매입단가", "D", 12),
        ("매입금액", "E", 14),
        ("현재가",   "F", 12),
        ("전일대비", "G", 12),
        ("등락률",   "H", 10),
        ("평가금액", "I", 14),
        ("평가손익", "J", 14),
        ("수익률",   "K", 10),
        ("비중(%)",  "L",  9),
    ]
    for text, col, width in cols:
        header_cell(ws[f"{col}5"], text)
        ws.column_dimensions[col].width = width
    ws.row_dimensions[5].height = 22

    # ── 데이터 행 ──
    holdings = portfolio_data.get("holdings", [])
    for i, item in enumerate(holdings):
        row = 6 + i
        ws.row_dimensions[row].height = 20
        ri = i  # row_idx for alternating color

        data_cell(ws.cell(row,  1), item.get("code", ""),        align="center", row_idx=ri)
        data_cell(ws.cell(row,  2), item.get("name", ""),        align="left",   row_idx=ri)
        data_cell(ws.cell(row,  3), item.get("quantity", 0),     fmt="#,##0",    align="right", row_idx=ri)
        data_cell(ws.cell(row,  4), item.get("avg_price", 0),    fmt="#,##0",    align="right", row_idx=ri)
        data_cell(ws.cell(row,  5), item.get("buy_amount", 0),   fmt="#,##0",    align="right", row_idx=ri)
        data_cell(ws.cell(row,  6), item.get("current_price", 0),fmt="#,##0",    align="right", row_idx=ri)

        change = item.get("change", 0)
        cc = ws.cell(row, 7)
        data_cell(cc, change, fmt='+#,##0;-#,##0;0', align="right", row_idx=ri)
        color_pnl(cc, change)

        chrate = item.get("change_rate", 0)
        rc = ws.cell(row, 8)
        data_cell(rc, chrate / 100 if chrate else 0, fmt='+0.00%;-0.00%;0.00%',
                  align="right", row_idx=ri)
        color_pnl(rc, chrate)

        data_cell(ws.cell(row, 9), item.get("eval_amount", 0), fmt="#,##0", align="right", row_idx=ri)

        pl = item.get("profit_loss", 0)
        plc = ws.cell(row, 10)
        data_cell(plc, pl, fmt='+#,##0;-#,##0;0', align="right", row_idx=ri)
        color_pnl(plc, pl)

        pr_val = item.get("profit_rate", 0)
        prc = ws.cell(row, 11)
        data_cell(prc, pr_val / 100 if pr_val else 0, fmt='+0.00%;-0.00%;0.00%',
                  align="right", row_idx=ri)
        color_pnl(prc, pr_val)

        # 비중 (eval / total_eval)
        wt = (item.get("eval_amount", 0) / total_eval * 100) if total_eval else 0
        data_cell(ws.cell(row, 12), round(wt, 1), fmt="0.0", align="right", row_idx=ri)

    # ── 합계 행 ──
    if holdings:
        tr = 6 + len(holdings)
        ws.row_dimensions[tr].height = 24
        ws.merge_cells(f"A{tr}:E{tr}")
        tc = ws[f"A{tr}"]
        tc.value     = "합  계"
        tc.font      = Font(name="맑은 고딕", bold=True, size=11)
        tc.fill      = PatternFill("solid", fgColor=C_TOTAL_BG)
        tc.alignment = Alignment(horizontal="center", vertical="center")
        tc.border    = thin_border("AAAAAA")

        for col_idx, val, fmt, sign in [
            (9,  total_eval,        '#,##0"원"',                    False),
            (10, total_profit,      '+#,##0;-#,##0;0"원"',          True),
            (11, total_prate / 100 if total_prate else 0,
                                    '+0.00%;-0.00%;0.00%',          True),
        ]:
            c = ws.cell(tr, col_idx)
            c.value          = val
            c.number_format  = fmt
            color = (C_UP if val >= 0 else C_DOWN) if sign else "1F3864"
            c.font           = Font(name="맑은 고딕", bold=True, size=11, color=color)
            c.fill           = PatternFill("solid", fgColor=C_TOTAL_BG)
            c.alignment      = Alignment(horizontal="right", vertical="center")
            c.border         = thin_border("AAAAAA")
        ws.cell(tr, 12).fill = PatternFill("solid", fgColor=C_TOTAL_BG)

    ws.freeze_panes = "A6"


def build_search_sheet(ws):
    ws.title = "종목 검색"

    ws.merge_cells("A1:I1")
    c = ws["A1"]
    c.value     = "🔍  종목 검색  —  종목코드(예: 005930) 또는 종목명(예: 삼성전자) 입력"
    c.font      = Font(name="맑은 고딕", bold=True, size=13, color="1F3864")
    c.fill      = PatternFill("solid", fgColor=C_SECTION_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # 입력 행
    ws["A3"].value     = "검색어 입력"
    ws["A3"].font      = Font(name="맑은 고딕", bold=True, size=11)
    ws["A3"].fill      = PatternFill("solid", fgColor=C_SECTION_BG)
    ws["A3"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A3"].border    = thin_border()
    ws.column_dimensions["A"].width = 14

    ws.merge_cells("B3:C3")
    ic = ws["B3"]
    ic.value     = ""
    ic.font      = Font(name="맑은 고딕", size=12)
    ic.fill      = PatternFill("solid", fgColor=C_INPUT_BG)
    ic.alignment = Alignment(horizontal="center", vertical="center")
    ic.border    = thin_border("1F3864")

    ws.merge_cells("D3:I3")
    nc = ws["D3"]
    nc.value     = "※ 입력 후 파일 저장 → python stock_manager.py 실행 → 결과 확인"
    nc.font      = Font(name="맑은 고딕", size=9, color="888888", italic=True)
    nc.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[3].height = 24
    ws.row_dimensions[2].height = 8
    ws.row_dimensions[4].height = 8

    # 결과 헤더 (행5)
    search_headers = [
        ("종목코드", "A",  9), ("종목명",   "B", 18), ("현재가",   "C", 12),
        ("전일대비", "D", 12), ("등락률",   "E", 10), ("시가",     "F", 12),
        ("고가",     "G", 12), ("저가",     "H", 12), ("거래량",   "I", 14),
    ]
    for text, col, width in search_headers:
        header_cell(ws[f"{col}5"], text)
        ws.column_dimensions[col].width = width
    ws.row_dimensions[5].height = 22

    # 빈 결과 행 (10행)
    for i in range(10):
        row = 6 + i
        ws.row_dimensions[row].height = 20
        for col in range(1, 10):
            data_cell(ws.cell(row, col), "", row_idx=i)


def update_search_results(ws, results):
    """검색 결과를 종목검색 시트에 기입"""
    for i in range(10):
        row = 6 + i
        for col in range(1, 10):
            ws.cell(row, col).value = None

    for i, item in enumerate(results[:10]):
        row = 6 + i
        data_cell(ws.cell(row, 1), item.get("code", ""),         align="center", row_idx=i)
        data_cell(ws.cell(row, 2), item.get("name", ""),         align="left",   row_idx=i)
        data_cell(ws.cell(row, 3), item.get("current_price", 0), fmt="#,##0",    align="right", row_idx=i)

        chg = item.get("change", 0)
        cc  = ws.cell(row, 4)
        data_cell(cc, chg, fmt='+#,##0;-#,##0;0', align="right", row_idx=i)
        color_pnl(cc, chg)

        cr = item.get("change_rate", 0)
        rc = ws.cell(row, 5)
        data_cell(rc, cr / 100 if cr else 0, fmt='+0.00%;-0.00%;0.00%', align="right", row_idx=i)
        color_pnl(rc, cr)

        data_cell(ws.cell(row, 6), item.get("open",   0), fmt="#,##0", align="right", row_idx=i)
        data_cell(ws.cell(row, 7), item.get("high",   0), fmt="#,##0", align="right", row_idx=i)
        data_cell(ws.cell(row, 8), item.get("low",    0), fmt="#,##0", align="right", row_idx=i)
        data_cell(ws.cell(row, 9), item.get("volume", 0), fmt="#,##0", align="right", row_idx=i)


def build_manual_input_sheet(ws):
    ws.title = "보유종목 입력"

    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value     = "📝  보유 종목 수동 입력  (KIS API 미사용 시)"
    c.font      = Font(name="맑은 고딕", bold=True, size=13, color="1F3864")
    c.fill      = PatternFill("solid", fgColor=C_SECTION_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:F2")
    nc = ws["A2"]
    nc.value     = "※ 종목코드(6자리)·보유수량·매입단가를 입력하세요. 종목명/매입금액은 자동 계산됩니다."
    nc.font      = Font(name="맑은 고딕", size=9, color="666666", italic=True)
    nc.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 6

    headers = [
        ("종목코드", "A", 12), ("보유수량", "B", 11),
        ("매입단가", "C", 13), ("비고",     "D", 20),
    ]
    for text, col, width in headers:
        header_cell(ws[f"{col}4"], text)
        ws.column_dimensions[col].width = width
    ws.row_dimensions[4].height = 22

    # 샘플 3행
    samples = [("005930", 10, 72000, ""), ("000660", 5, 140000, ""), ("035720", 8, 53000, "")]
    for i, (code, qty, price, note) in enumerate(samples):
        row = 5 + i
        data_cell(ws.cell(row, 1), code,  align="center", row_idx=i)
        data_cell(ws.cell(row, 2), qty,   fmt="#,##0", align="right",  row_idx=i)
        data_cell(ws.cell(row, 3), price, fmt="#,##0", align="right",  row_idx=i)
        data_cell(ws.cell(row, 4), note,  align="left",   row_idx=i)
        ws.row_dimensions[row].height = 20

    # 빈 입력 행 20개
    for i in range(20):
        row = 8 + i
        for col in range(1, 5):
            c = ws.cell(row, col)
            c.fill   = PatternFill("solid", fgColor=C_ROW_ODD if i % 2 == 0 else C_ROW_EVEN)
            c.border = thin_border()
            c.font   = Font(name="맑은 고딕", size=10)
            if col in (2, 3):
                c.number_format = "#,##0"
                c.alignment = Alignment(horizontal="right", vertical="center")
            else:
                c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 20

    ws.freeze_panes = "A5"


def build_settings_sheet(ws, config):
    ws.title = "설정"

    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value     = "⚙️  API 설정  /  사용 방법"
    c.font      = Font(name="맑은 고딕", bold=True, size=13, color="1F3864")
    c.fill      = PatternFill("solid", fgColor=C_SECTION_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 42
    ws.column_dimensions["D"].width = 35

    def s_label(row, text):
        c = ws.cell(row, 2)
        c.value     = text
        c.font      = Font(name="맑은 고딕", bold=True, size=10)
        c.fill      = PatternFill("solid", fgColor=C_SECTION_BG)
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.border    = thin_border("BBBBBB")
        ws.row_dimensions[row].height = 22

    def s_value(row, val, note=""):
        c = ws.cell(row, 3)
        c.value     = str(val)
        c.font      = Font(name="맑은 고딕", size=10)
        c.fill      = PatternFill("solid", fgColor="FFFFFF")
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border    = thin_border("BBBBBB")
        if note:
            n = ws.cell(row, 4)
            n.value     = note
            n.font      = Font(name="맑은 고딕", size=9, color="888888", italic=True)
            n.alignment = Alignment(horizontal="left", vertical="center")

    kis = config.get("kis", {})
    s_label(3, "API 제공자");    s_value(3, config.get("api_provider","pykrx"), '"pykrx" 또는 "kis"')
    s_label(4, "KIS APP_KEY");   s_value(4, kis.get("app_key",""), "한국투자증권 OpenAPI 포털에서 발급")
    s_label(5, "KIS APP_SECRET");s_value(5, kis.get("app_secret",""))
    s_label(6, "계좌번호");       s_value(6, kis.get("account_no",""), '예) "50123456-01"')
    s_label(7, "모의투자 여부");  s_value(7, str(kis.get("is_paper",False)), "모의투자: True  /  실전: False")

    ws.row_dimensions[8].height = 10

    # 사용 안내
    ws.merge_cells("B9:D9")
    h = ws["B9"]
    h.value     = "📌  사용 방법"
    h.font      = Font(name="맑은 고딕", bold=True, size=11)
    h.fill      = PatternFill("solid", fgColor=C_SECTION_BG)
    h.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[9].height = 22

    guide = (
        "【 pykrx 모드 (기본값, 무료) 】\n"
        "  1. pip install pykrx openpyxl requests 설치\n"
        "  2. '보유종목 입력' 시트에 종목코드·수량·매입단가 입력 후 저장\n"
        "  3. python stock_manager.py 실행 → 포트폴리오 자동 계산\n"
        "  4. 종목검색: '종목 검색' 시트 B3셀에 코드 또는 이름 입력 후 실행\n\n"
        "【 KIS API 모드 (실시간, 잔고 자동 연동) 】\n"
        "  1. https://apiportal.koreainvestment.com 에서 앱 등록\n"
        "  2. APP_KEY, APP_SECRET, 계좌번호를 config.json 에 입력\n"
        "  3. api_provider 를 \"kis\" 로 변경 후 저장\n"
        "  4. python stock_manager.py 실행\n\n"
        "【 자동 새로고침 】\n"
        "  refresh.bat 더블클릭 또는 작업 스케줄러에 등록하면 자동 업데이트 됩니다."
    )
    ws.merge_cells("B10:D22")
    gc = ws["B10"]
    gc.value          = guide
    gc.font           = Font(name="맑은 고딕", size=10)
    gc.fill           = PatternFill("solid", fgColor=C_NOTE_BG)
    gc.alignment      = Alignment(horizontal="left", vertical="top", wrap_text=True)
    gc.border         = thin_border("CCCCCC")
    ws.row_dimensions[10].height = 180


# ============================================================
#  엑셀 파일 열기/생성 헬퍼
# ============================================================
def open_or_create_workbook(path):
    if os.path.exists(path):
        wb = openpyxl.load_workbook(path)
    else:
        wb = openpyxl.Workbook()
        default = wb.active
        if default:
            wb.remove(default)
    return wb


def get_or_replace_sheet(wb, title, position):
    """기존 시트 삭제 후 새로 생성"""
    if title in wb.sheetnames:
        del wb[title]
    ws = wb.create_sheet(title, position)
    return ws


# ============================================================
#  포트폴리오 데이터 수집
# ============================================================
def fetch_portfolio_pykrx(manual_list, pykrx_prov):
    """pykrx 로 수동 포트폴리오 가격 조회"""
    holdings = []
    total_eval = total_buy = 0

    for item in manual_list:
        code     = str(item.get("code", "")).zfill(6)
        qty      = int(item.get("quantity", 0))
        avg_p    = float(item.get("avg_price", 0))
        if qty <= 0:
            continue

        info = pykrx_prov.get_stock_price(code)
        if not info:
            print(f"  ⚠  {code} 조회 실패, 건너뜁니다.")
            continue

        current    = info["current_price"]
        eval_amt   = current * qty
        buy_amt    = int(avg_p * qty)
        pl         = eval_amt - buy_amt
        pr         = pl / buy_amt * 100 if buy_amt else 0.0

        holdings.append({
            "code":          code,
            "name":          info.get("name", code),
            "quantity":      qty,
            "avg_price":     avg_p,
            "current_price": current,
            "change":        info.get("change", 0),
            "change_rate":   info.get("change_rate", 0.0),
            "eval_amount":   eval_amt,
            "buy_amount":    buy_amt,
            "profit_loss":   pl,
            "profit_rate":   round(pr, 2),
        })
        total_eval += eval_amt
        total_buy  += buy_amt
        print(f"  ✓  {info.get('name', code):12s}  현재가 {current:>9,}원  "
              f"{'▲' if info.get('change',0)>=0 else '▼'}{abs(info.get('change_rate',0)):.2f}%")

    total_profit = total_eval - total_buy
    summary = {
        "total_eval":        total_eval,
        "total_buy":         total_buy,
        "total_profit":      total_profit,
        "total_profit_rate": (total_profit / total_buy * 100) if total_buy else 0.0,
    }
    return holdings, summary


# ============================================================
#  엑셀에서 수동 입력 읽기
# ============================================================
def read_manual_portfolio(wb):
    ws = None
    for name in wb.sheetnames:
        if "보유종목" in name or "입력" in name:
            ws = wb[name]
            break
    if not ws:
        return []

    result = []
    for row in ws.iter_rows(min_row=5, max_row=50, values_only=True):
        code = str(row[0]).strip() if row[0] else ""
        if not code or code in ("None", "") or len(code) < 5:
            continue
        try:
            qty   = int(float(str(row[1]).replace(",", ""))) if row[1] else 0
            price = float(str(row[2]).replace(",", ""))      if row[2] else 0
        except ValueError:
            continue
        if qty > 0:
            result.append({"code": code.zfill(6), "quantity": qty, "avg_price": price})
    return result


def read_search_query(wb):
    for name in wb.sheetnames:
        if "검색" in name:
            ws = wb[name]
            val = ws["B3"].value
            return str(val).strip() if val else ""
    return ""


# ============================================================
#  메인
# ============================================================
def main():
    print("=" * 56)
    print("   [주식 포트폴리오 매니저]")
    print("=" * 56)

    # 설정 로드
    if not os.path.exists(CONFIG_FILE):
        with open(CONFIG_FILE, "w", encoding="utf-8") as f:
            json.dump(DEFAULT_CONFIG, f, ensure_ascii=False, indent=2)
        print(f"✅ 설정 파일 생성 → {CONFIG_FILE}\n   config.json 을 수정해 API 키를 입력하세요.\n")

    with open(CONFIG_FILE, encoding="utf-8") as f:
        config = json.load(f)

    provider = config.get("api_provider", "pykrx")
    print(f"  API 제공자: {provider}\n")

    # ── API 초기화 ──
    api_kis   = None
    api_pykrx = None

    if provider == "kis":
        kc = config.get("kis", {})
        if kc.get("app_key", "").startswith("YOUR"):
            print("  ⚠  KIS APP_KEY 미설정 → pykrx 모드로 전환")
        else:
            api_kis = KISApi(kc["app_key"], kc["app_secret"],
                             kc["account_no"], kc.get("is_paper", False))
            print("  [KIS] 인증 중 ...")
            if not api_kis.authenticate():
                print("  ⚠  KIS 인증 실패 → pykrx 모드로 전환")
                api_kis = None

    if api_kis is None:
        # pykrx 우선, 없으면 네이버 금융 API 폴백
        _pykrx = PykrxProvider()
        if _pykrx.available:
            api_pykrx = _pykrx
            print("  [pykrx] 사용")
        else:
            api_pykrx = NaverStockProvider()
            print("  [Naver] 네이버 금융 API 사용 (pykrx 미설치)")

    # ── 엑셀 파일 열기 ──
    wb = open_or_create_workbook(EXCEL_FILE)

    # ── 포트폴리오 데이터 ──
    holdings, summary = [], {}

    if api_kis:
        print("\n  [KIS] 잔고 조회 중 ...")
        balance = api_kis.get_balance()
        if balance:
            for item in balance["holdings"]:
                pi = api_kis.get_stock_price(item["code"])
                if pi:
                    item["change"]      = pi.get("change", 0)
                    item["change_rate"] = pi.get("change_rate", 0.0)
                print(f"  OK  {item.get('name',''):12s}  현재가 {item.get('current_price',0):>9,}원")
            holdings = balance["holdings"]
            summary  = {k: balance[k] for k in
                        ("total_eval","total_buy","total_profit","total_profit_rate") if k in balance}
            summary["deposit"]     = balance.get("deposit", 0)
            summary["total_asset"] = balance.get("total_asset", 0)
            print(f"\n  총 {len(holdings)}개 종목 조회 완료")
        else:
            print("  [!] 잔고 조회 실패")

    if not holdings and api_pykrx:
        # 수동 포트폴리오 우선 (엑셀 → config 순)
        manual = read_manual_portfolio(wb) or config.get("portfolio", [])
        if not manual:
            print("\n  [i] '보유종목 입력' 시트 또는 config.json 에 종목을 입력하세요.")
        else:
            provider_name = "pykrx" if isinstance(api_pykrx, PykrxProvider) else "Naver"
            print(f"\n  [{provider_name}] {len(manual)}개 종목 가격 조회 중 ...")
            holdings, summary = fetch_portfolio_pykrx(manual, api_pykrx)

    # ── 종목 검색 ──
    search_query   = read_search_query(wb)
    search_results = []

    if search_query:
        print(f"\n  [검색] '{search_query}' ...")
        if api_kis:
            results = api_kis.search_stock(search_query)
            for r in results[:5]:
                pi = api_kis.get_stock_price(r["code"])
                if pi:
                    search_results.append(pi)
        elif api_pykrx:
            if search_query.isdigit():
                pi = api_pykrx.get_stock_price(search_query.zfill(6))
                if pi:
                    search_results.append(pi)
            else:
                matches = api_pykrx.search_by_name(search_query)
                for m in matches[:5]:
                    pi = api_pykrx.get_stock_price(m["code"])
                    if pi:
                        search_results.append(pi)
        print(f"  → {len(search_results)}건 결과")

    # ── 시트 생성/업데이트 ──
    print("\n  엑셀 파일 생성 중 ...")

    ws_port = get_or_replace_sheet(wb, "포트폴리오", 0)
    build_portfolio_sheet(ws_port, {"holdings": holdings, "summary": summary})

    if "종목 검색" not in wb.sheetnames:
        ws_search = get_or_replace_sheet(wb, "종목 검색", 1)
        build_search_sheet(ws_search)
    else:
        ws_search = wb["종목 검색"]

    if search_results:
        update_search_results(ws_search, search_results)

    if "보유종목 입력" not in wb.sheetnames:
        ws_manual = get_or_replace_sheet(wb, "보유종목 입력", 2)
        build_manual_input_sheet(ws_manual)

    if "설정" not in wb.sheetnames:
        ws_settings = get_or_replace_sheet(wb, "설정", 3)
        build_settings_sheet(ws_settings, config)

    wb.active = wb["포트폴리오"]

    # 저장
    wb.save(EXCEL_FILE)
    print(f"\n✅ 저장 완료: {EXCEL_FILE}")

    # Windows에서 자동으로 엑셀 열기
    try:
        os.startfile(os.path.abspath(EXCEL_FILE))
    except Exception:
        pass


if __name__ == "__main__":
    main()
